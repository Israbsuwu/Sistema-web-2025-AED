from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, update_session_auth_hash
from django.contrib.auth.models import User
from .models import Cliente, Categoria, Producto, Proveedor, Venta, DetalleVenta, Comentario, Carrito, CarritoDetalle, Stock
from .forms import LoginForm, RegisterForm, ProductoForm, ProveedorForm, CategoriaForm, StockForm
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import logout as auth_logout
from django.contrib import messages
from django.contrib.auth.models import User
from .models import Cliente, Producto
from django.db.models import Q, Sum, Count
from django.db.models.functions import Coalesce
from decimal import Decimal, InvalidOperation
from django.http import JsonResponse
import json
import openpyxl
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
import os
from django.conf import settings
from django.contrib.staticfiles import finders

# Create your views here.
# INICIO DE SESIÓN
def index_view(request):
    error = None
    initial = {} if request.method != 'POST' else request.POST
    form = LoginForm(initial if request.method == 'POST' else None)
    if request.method == 'POST':
        if form.is_valid():
            user_auth = form.cleaned_data['user_auth']
            from django.contrib.auth import login
            login(request, user_auth)
            # Funcionalidad de 'Recuérdame'
            if request.POST.get('remember'):
                request.session.set_expiry(1209600)  # 2 semanas
            else:
                request.session.set_expiry(0)  # hasta cerrar navegador
            if user_auth.is_staff or user_auth.is_superuser:
                return redirect('admin_dashboard')
            return redirect('catalogo_home')
        else:
            error_list = [e for e in form.errors.values()]
            error = error_list[0][0] if error_list and isinstance(error_list[0], list) else str(form.errors)
    else:
        # Limpiar campos y mensajes al volver al login
        form = LoginForm()
        error = None
    return render(request, 'vlastef_app/index.html', {'form': form, 'error': error})

# REGISTRO
def register_view(request):
    error = None
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user, cliente = form.save()
            from django.contrib import messages
            messages.success(request, 'Registro exitoso. Ahora puedes iniciar sesión con tu usuario y contraseña.')
            return redirect('index')
        else:
            # Mostrar solo el primer error general o de campo
            error_list = [e for e in form.errors.values()]
            error = error_list[0][0] if error_list and isinstance(error_list[0], list) else str(form.errors)
    else:
        form = RegisterForm()
        error = None
    return render(request, 'vlastef_app/register.html', {'form': form, 'error': error})

# HOME

# --- CATÁLOGO ---
from django.db.models import Q

def catalogo_home_view(request):
    # Filtros por género y categoría
    genero = request.GET.get('genero')
    categoria = request.GET.get('categoria')
    precio_min = request.GET.get('precio_min')
    precio_max = request.GET.get('precio_max')
    talla = request.GET.get('talla')
    color = request.GET.get('color')
    q = request.GET.get('q')

    productos = Producto.objects.all()
    
    # Búsqueda por texto
    if q:
        productos = productos.filter(
            Q(nombre__icontains=q) | 
            Q(descripcion__icontains=q) |
            Q(categoria__nombre__icontains=q)
        )

    # Filtrado por género: Hombre, Mujer, Ambos
    if genero == 'H':
        productos = productos.filter(genero__in=['H', 'A'])
    elif genero == 'M':
        productos = productos.filter(genero__in=['M', 'A'])
    
    # Filtrado por categoría (si se selecciona una categoría específica)
    if categoria:
        productos = productos.filter(categoria_id=categoria)
        
    # Filtros avanzados
    if precio_min:
        try:
            # Reemplazar coma por punto para decimales y limpiar espacios
            val_min = precio_min.replace(',', '.').strip()
            productos = productos.filter(precio_venta__gte=Decimal(val_min))
        except (ValueError, InvalidOperation):
            pass
            
    if precio_max:
        try:
            val_max = precio_max.replace(',', '.').strip()
            productos = productos.filter(precio_venta__lte=Decimal(val_max))
        except (ValueError, InvalidOperation):
            pass
            
    if talla:
        productos = productos.filter(tallas__icontains=talla.strip())
        
    if color:
        productos = productos.filter(colores__icontains=color.strip())

    # Ordenamiento
    orden = request.GET.get('orden')
    
    if orden == 'price_asc':
        productos = productos.order_by('precio_venta')
    elif orden == 'price_desc':
        productos = productos.order_by('-precio_venta')
    elif orden == 'new':
        productos = productos.order_by('-id')
    elif orden == 'popular':
        # Ordenar por cantidad total vendida
        productos = productos.annotate(
            total_vendido=Coalesce(Sum('detalleventa__cantidad'), 0)
        ).order_by('-total_vendido')
    else:
        productos = productos.order_by('-id')

    categorias = Categoria.objects.all()
    return render(request, 'vlastef_app/home.html', {
        'productos': productos,
        'categorias': categorias,
        'genero': genero,
        'categoria': categoria,
        # Pasar valores actuales para mantenerlos en los inputs (opcional, pero buena UX)
        'precio_min': precio_min,
        'precio_max': precio_max,
        'talla': talla,
        'color': color,
        'orden': orden,
    })

# @login_required(login_url='index')
# def catalogo_carrito_view(request):
#     # Deprecated
#     return redirect('catalogo_home')

@login_required(login_url='index')
def add_to_cart_view(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            prod_id = data.get('id')
            cantidad = int(data.get('qty', 1))
            talla = data.get('talla')
            color = data.get('color')
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Datos inválidos'}, status=400)

        if not prod_id:
            return JsonResponse({'error': 'Producto no especificado'}, status=400)

        producto = get_object_or_404(Producto, id=prod_id)
        
        try:
            cliente = request.user.cliente
        except Cliente.DoesNotExist:
             return JsonResponse({'error': 'Usuario no es cliente'}, status=403)

        # Singleton Cart Logic
        carrito = Carrito.objects.filter(cliente=cliente, activo=True).first()
        if not carrito:
            carrito = Carrito.objects.create(cliente=cliente, activo=True)
        
        detalle, created = CarritoDetalle.objects.get_or_create(
            carrito=carrito,
            producto=producto,
            talla=talla,
            color=color,
            defaults={'cantidad': 0, 'subtotal': 0}
        )
        
        detalle.cantidad += cantidad
        detalle.subtotal = detalle.cantidad * producto.precio_venta
        detalle.save()
        
        # Recalcular total carrito items
        total_items = carrito.detalles.aggregate(total=Sum('cantidad'))['total'] or 0
        
        return JsonResponse({'success': True, 'total_items': total_items, 'message': 'Producto agregado'})
    return JsonResponse({'error': 'Método no permitido'}, status=405)

@login_required(login_url='index')
def update_cart_view(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            item_id = data.get('id')
            cantidad = int(data.get('qty', 1))
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Datos inválidos'}, status=400)

        try:
            cliente = request.user.cliente
        except Cliente.DoesNotExist:
             return JsonResponse({'error': 'Usuario no es cliente'}, status=403)

        carrito = get_object_or_404(Carrito, cliente=cliente, activo=True)
        detalle = get_object_or_404(CarritoDetalle, carrito=carrito, id=item_id)
        
        if cantidad > 0:
            detalle.cantidad = cantidad
            detalle.subtotal = detalle.cantidad * detalle.producto.precio_venta
            detalle.save()
        else:
            detalle.delete()
            
        # Recalcular totales
        detalles = carrito.detalles.all()
        total = sum(item.subtotal for item in detalles)
        total_items = sum(item.cantidad for item in detalles)
        
        return JsonResponse({
            'success': True, 
            'total': float(total), 
            'subtotal': float(detalle.subtotal) if cantidad > 0 else 0,
            'total_items': total_items
        })
    return JsonResponse({'error': 'Método no permitido'}, status=405)

@login_required(login_url='index')
def remove_from_cart_view(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            item_id = data.get('id')
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Datos inválidos'}, status=400)

        try:
            cliente = request.user.cliente
        except Cliente.DoesNotExist:
             return JsonResponse({'error': 'Usuario no es cliente'}, status=403)

        carrito = get_object_or_404(Carrito, cliente=cliente, activo=True)
        detalle = get_object_or_404(CarritoDetalle, carrito=carrito, id=item_id)
        detalle.delete()
        
        # Recalcular totales
        detalles = carrito.detalles.all()
        total = sum(item.subtotal for item in detalles)
        total_items = sum(item.cantidad for item in detalles)
        
        return JsonResponse({'success': True, 'total': float(total), 'total_items': total_items})
    return JsonResponse({'error': 'Método no permitido'}, status=405)

@login_required(login_url='index')
def clear_cart_view(request):
    if request.method == 'POST':
        try:
            cliente = request.user.cliente
            carrito = get_object_or_404(Carrito, cliente=cliente, activo=True)
            carrito.detalles.all().delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    return JsonResponse({'error': 'Método no permitido'}, status=405)

@login_required(login_url='index')
def get_cart_data_view(request):
    try:
        cliente = request.user.cliente
        # Singleton Cart: Ensure only one active cart exists
        carrito = Carrito.objects.filter(cliente=cliente, activo=True).first()
        if not carrito:
            carrito = Carrito.objects.create(cliente=cliente, activo=True)
        
        # Clean up duplicates if any (paranoid check)
        duplicates = Carrito.objects.filter(cliente=cliente, activo=True).exclude(id=carrito.id)
        if duplicates.exists():
            duplicates.delete()

        detalles = carrito.detalles.select_related('producto').all()
        
        items = []
        for d in detalles:
            items.append({
                'id': d.id, # ID del detalle, no del producto
                'product_id': d.producto.id,
                'title': d.producto.nombre,
                'price': float(d.producto.precio_venta),
                'qty': d.cantidad,
                'img': d.producto.imagen.url if d.producto.imagen else None,
                'subtotal': float(d.subtotal)
            })
            
        total = sum(item['subtotal'] for item in items)
        
        return JsonResponse({
            'success': True,
            'items': items,
            'total': total,
            'count': sum(item['qty'] for item in items)
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def catalogo_ver_producto_view(request, prod_id):
    producto = get_object_or_404(Producto, id=prod_id)
    comentarios = Comentario.objects.filter(producto=producto).select_related('cliente').order_by('-fecha')
    
    user_comment = None
    
    if request.user.is_authenticated:
        try:
            cliente = request.user.cliente
            user_comment = Comentario.objects.filter(cliente=cliente, producto=producto).first()
        except Cliente.DoesNotExist:
            pass

    if request.method == 'POST' and request.user.is_authenticated:
        try:
            cliente = request.user.cliente
            texto = request.POST.get('texto')
            calificacion = int(request.POST.get('calificacion', 5))
            
            if user_comment:
                user_comment.texto = texto
                user_comment.calificacion = calificacion
                user_comment.save()
                messages.success(request, 'Tu comentario ha sido actualizado.')
            else:
                Comentario.objects.create(
                    cliente=cliente,
                    producto=producto,
                    texto=texto,
                    calificacion=calificacion
                )
                messages.success(request, 'Gracias por tu comentario.')
            return redirect('catalogo_ver_producto', prod_id=prod_id)
        except Cliente.DoesNotExist:
            messages.error(request, 'Debes ser un cliente registrado para comentar.')
        except Exception as e:
            messages.error(request, f'Error al guardar comentario: {e}')

    return render(request, 'vlastef_app/Ver_Producto.html', {
        'producto': producto,
        'comentarios': comentarios,
        'user_comment': user_comment
    })

@login_required(login_url='index')
def cliente_editar_perfil_view(request):
    cliente = get_object_or_404(Cliente, usuario=request.user)
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        
        # Validar email único
        if email:
            if User.objects.filter(email=email).exclude(id=request.user.id).exists():
                messages.error(request, 'Ya existe una cuenta con ese correo.')
                return redirect('catalogo_home')
        
        # Validar nombres y apellidos no vacíos
        nombres = request.POST.get('nombres', '').strip()
        apellidos = request.POST.get('apellidos', '').strip()
        if not nombres or not apellidos:
            messages.error(request, 'El nombre y el apellido son obligatorios.')
            return redirect('catalogo_home')

        # Validar teléfono no vacío
        if not telefono:
            messages.error(request, 'El número de teléfono no puede quedar vacío.')
            return redirect('catalogo_home')

        # Validar teléfono único y formato
        import re
        if not re.match(r'^[578]\d{7}$', telefono):
            messages.error(request, 'Por favor, ingrese un número válido.')
            return redirect('catalogo_home')

        if Cliente.objects.filter(telefono=telefono).exclude(id=cliente.id).exists():
            messages.error(request, 'Ya existe una cuenta con ese número de teléfono.')
            return redirect('catalogo_home')
        
        # Validar fecha de nacimiento
        fecha_nacimiento = request.POST.get('fecha_nacimiento')
        if fecha_nacimiento:
            from datetime import date, datetime
            try:
                fecha_obj = datetime.strptime(fecha_nacimiento, '%Y-%m-%d').date()
                if fecha_obj > date.today():
                     messages.error(request, 'Fecha de nacimiento inválida.')
                     return redirect('catalogo_home')
                if fecha_obj.year < 1900:
                     messages.error(request, 'Fecha de nacimiento inválida.')
                     return redirect('catalogo_home')
            except ValueError:
                pass

        # Actualizar datos de User
        user = request.user
        user.email = email
        user.first_name = request.POST.get('nombres', '')
        user.last_name = request.POST.get('apellidos', '')

        # Cambio de contraseña
        current_password = request.POST.get('current_password')
        new_password = request.POST.get('new_password')
        password_changed = False
        
        if current_password and new_password:
            if not user.check_password(current_password):
                messages.error(request, 'La contraseña actual es incorrecta.')
                return redirect('catalogo_home')
            
            if len(new_password) < 8:
                messages.error(request, 'La nueva contraseña debe tener al menos 8 caracteres.')
                return redirect('catalogo_home')
                
            user.set_password(new_password)
            password_changed = True

        user.save()

        if password_changed:
            update_session_auth_hash(request, user)
        
        # Actualizar datos de Cliente
        cliente.nombres = request.POST.get('nombres', '')
        cliente.apellidos = request.POST.get('apellidos', '')
        cliente.telefono = telefono
        cliente.direccion = request.POST.get('direccion')
        cliente.sexo = request.POST.get('genero')
        cliente.fecha_nacimiento = request.POST.get('fecha_nacimiento') or None
        
        if 'foto' in request.FILES:
            # Eliminar foto anterior si existe
            if cliente.foto:
                import os
                from django.conf import settings
                try:
                    old_path = os.path.join(settings.MEDIA_ROOT, cliente.foto.name)
                    if os.path.exists(old_path):
                        os.remove(old_path)
                except Exception:
                    pass
            cliente.foto = request.FILES['foto']
            
        cliente.save()
        
        messages.success(request, 'Perfil actualizado correctamente.')
        return redirect('catalogo_home')
    
    return redirect('catalogo_home')



# CERRAR SESIÓN
def logout_view(request):
    # Limpiar mensajes pendientes antes de cerrar sesión
    storage = messages.get_messages(request)
    for _ in storage:
        pass
    auth_logout(request)
    return redirect('index')


# VERIFICA SI ES STAFF O SUPERUSUARIO
def staff_required(user):
    return user.is_active and (user.is_staff or user.is_superuser)


# PANEL DE ADMINISTRACIÓN
# DASHBOARD
@user_passes_test(staff_required, login_url='index')
def admin_dashboard_view(request):
    ultimas_ventas = []
    productos_mas_vendidos = []
    productos_recientes = []
    return render(request, 'vlastef_app/admin_panel/dashboard.html', {
        'user': request.user,
        'ultimas_ventas': ultimas_ventas,
        'productos_mas_vendidos': productos_mas_vendidos,
        'productos_recientes': productos_recientes,
    })


# CLIENTES
@user_passes_test(staff_required, login_url='index')
def admin_clientes_view(request):
    q = request.GET.get('q', '').strip()
    clientes_qs = Cliente.objects.select_related('usuario').order_by('-id')
    if q:
        clientes_qs = clientes_qs.filter(
            Q(usuario__username__icontains=q) |
            Q(usuario__email__icontains=q) |
            Q(nombres__icontains=q) |
            Q(apellidos__icontains=q) |
            Q(telefono__icontains=q)
        )
    
    total_clientes = clientes_qs.count()
    
    return render(request, 'vlastef_app/admin_panel/clientes.html', {
        'clientes': clientes_qs,
        'query': q,
        'total_clientes': total_clientes,
    })

# Detalles del cliente
@user_passes_test(staff_required, login_url='index')
def admin_cliente_detalle_view(request, user_id):
    cliente = get_object_or_404(Cliente, usuario_id=user_id)
    return render(request, 'vlastef_app/admin_panel/cliente_detalle.html', {'cliente': cliente})

# Editar campos del cliente
@user_passes_test(staff_required, login_url='index')
def admin_cliente_editar_view(request, user_id):
    cliente = get_object_or_404(Cliente, usuario_id=user_id)
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        
        # Validar email único
        if email:
            if User.objects.filter(email=email).exclude(id=user_id).exists():
                messages.error(request, 'Ya existe una cuenta con ese correo.')
                return render(request, 'vlastef_app/admin_panel/cliente_editar.html', {'cliente': cliente})
        
        # Validar teléfono único
        if telefono:
            if Cliente.objects.filter(telefono=telefono).exclude(id=cliente.id).exists():
                messages.error(request, 'Ya existe una cuenta con ese número de teléfono.')
                return render(request, 'vlastef_app/admin_panel/cliente_editar.html', {'cliente': cliente})
        
        # Actualizar datos de User
        user = cliente.usuario
        user.email = email
        user.first_name = request.POST.get('nombres') # Opcional si se usa nombres de Cliente
        user.last_name = request.POST.get('apellidos') # Opcional
        
        # Actualizar staff status
        is_staff = request.POST.get('is_staff') == 'on'
        user.is_staff = is_staff
        
        user.save()
        
        # Actualizar datos de Cliente
        cliente.nombres = request.POST.get('nombres')
        cliente.apellidos = request.POST.get('apellidos')
        cliente.telefono = telefono
        cliente.direccion = request.POST.get('direccion')
        cliente.save()
        
        messages.success(request, 'Cliente actualizado correctamente.')
        return redirect('admin_clientes')
    
    return render(request, 'vlastef_app/admin_panel/cliente_editar.html', {'cliente': cliente})

# Activar cliente
@user_passes_test(staff_required, login_url='index')
def admin_cliente_activar_view(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.is_active = True
    user.save()
    messages.success(request, f'Usuario {user.username} activado.')
    return redirect('admin_clientes')

# Desactivar cliente
@user_passes_test(staff_required, login_url='index')
def admin_cliente_desactivar_view(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.is_active = False
    user.save()
    messages.success(request, f'Usuario {user.username} desactivado.')
    return redirect('admin_clientes')

# Eliminar cliente
@user_passes_test(staff_required, login_url='index')
def admin_cliente_eliminar_view(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.delete() # Esto eliminará también al Cliente por el on_delete=CASCADE
    messages.success(request, 'Cliente eliminado correctamente.')
    return redirect('admin_clientes')

# Restablecer contraseña del cliente
@user_passes_test(staff_required, login_url='index')
def admin_cliente_reset_password_view(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        new_password = request.POST.get('new_password')
        if not new_password or len(new_password.strip()) < 8:
            messages.error(request, 'La contraseña debe tener al menos 8 caracteres.')
            return redirect('admin_cliente_editar', user_id=user.id)
        user.set_password(new_password)
        user.save()
        messages.success(request, f'Contraseña restablecida para {user.username}.')
        return redirect('admin_clientes')
    return redirect('admin_cliente_editar', user_id=user.id)


# CATEGORÍAS
@user_passes_test(staff_required, login_url='index')
def admin_categorias_view(request):
    q = request.GET.get('q', '').strip()
    categorias_qs = Categoria.objects.all().order_by('-id')
    if q:
        categorias_qs = categorias_qs.filter(
            Q(nombre__icontains=q) |
            Q(descripcion__icontains=q)
        )
    
    total_categorias = categorias_qs.count()
    
    return render(request, 'vlastef_app/admin_panel/categorias.html', {
        'categorias': categorias_qs,
        'query': q,
        'total_categorias': total_categorias,
    })

# Crear categoría
@user_passes_test(staff_required, login_url='index')
def admin_categoria_crear_view(request):
    if request.method == 'POST':
        form = CategoriaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoría creada exitosamente.')
            return redirect('admin_categorias')
        else:
            nombre_errors = form.errors.get('nombre')
            if nombre_errors and any('Ya existe una categoría con ese nombre.' in str(e) for e in nombre_errors):
                messages.error(request, 'Ya existe una categoría con ese nombre.')
            else:
                messages.error(request, 'Corrige los errores en el formulario.')
            # Mantener valores ingresados en el template actual
            categoria = Categoria(nombre=request.POST.get('nombre', ''), descripcion=request.POST.get('descripcion', ''))
            return render(request, 'vlastef_app/admin_panel/categoria_form.html', {'titulo': 'Nueva Categoría', 'categoria': categoria})
    return render(request, 'vlastef_app/admin_panel/categoria_form.html', {'titulo': 'Nueva Categoría'})

# Editar categoría
@user_passes_test(staff_required, login_url='index')
def admin_categoria_editar_view(request, cat_id):
    categoria = get_object_or_404(Categoria, id=cat_id)
    if request.method == 'POST':
        form = CategoriaForm(request.POST, instance=categoria)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoría actualizada exitosamente.')
            return redirect('admin_categorias')
        else:
            nombre_errors = form.errors.get('nombre')
            if nombre_errors and any('Ya existe otra categoría con ese nombre.' in str(e) for e in nombre_errors):
                messages.error(request, 'Ya existe otra categoría con ese nombre.')
            else:
                messages.error(request, 'Corrige los errores en el formulario.')
    return render(request, 'vlastef_app/admin_panel/categoria_form.html', {'categoria': categoria, 'titulo': 'Editar Categoría'})

# Eliminar categoría
@user_passes_test(staff_required, login_url='index')
def admin_categoria_eliminar_view(request, cat_id):
    categoria = get_object_or_404(Categoria, id=cat_id)
    categoria.delete()
    messages.success(request, 'Categoría eliminada exitosamente.')
    return redirect('admin_categorias')

# Detalles de la categoría
@user_passes_test(staff_required, login_url='index')
def admin_categoria_detalle_view(request, cat_id):
    categoria = get_object_or_404(Categoria.objects.prefetch_related('productos'), id=cat_id)
    productos = categoria.productos.all().select_related('categoria', 'proveedor')
    total_productos = productos.count()
    return render(request, 'vlastef_app/admin_panel/categoria_detalle.html', {
        'categoria': categoria,
        'productos': productos,
        'total_productos': total_productos,
    })


# PRODUCTOS
@user_passes_test(staff_required, login_url='index')
def admin_productos_view(request):
    q = request.GET.get('q', '').strip()
    productos_qs = Producto.objects.select_related('categoria', 'proveedor').order_by('-id')
    if q:
        productos_qs = productos_qs.filter(
            Q(nombre__icontains=q) |
            Q(descripcion__icontains=q) |
            Q(categoria__nombre__icontains=q) |
            Q(proveedor__nombre__icontains=q)
        )
    return render(request, 'vlastef_app/admin_panel/productos.html', {
        'productos': productos_qs,
        'query': q,
    })

# Crear producto
@user_passes_test(staff_required, login_url='index')
def admin_producto_crear_view(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            producto = form.save()
            # Reubicar imagen principal en carpeta por id
            if producto.imagen and producto.imagen.name:
                from django.conf import settings
                import os, shutil
                orig_rel = producto.imagen.name.replace('\\','/')
                if orig_rel.startswith('productos/') and not orig_rel.startswith(f'productos/{producto.id}/'):
                    orig_abs = os.path.join(settings.MEDIA_ROOT, orig_rel)
                    target_dir = os.path.join(settings.MEDIA_ROOT, 'productos', str(producto.id))
                    os.makedirs(target_dir, exist_ok=True)
                    new_name = os.path.basename(orig_abs)
                    try:
                        target_princ = os.path.join(target_dir, new_name)
                        if os.path.exists(target_princ):
                            try:
                                os.remove(target_princ)
                            except Exception:
                                pass
                        shutil.move(orig_abs, target_princ)
                        producto.imagen.name = f"productos/{producto.id}/{new_name}"
                        producto.save(update_fields=['imagen'])
                    except Exception:
                        pass
            # Guardar imágenes adicionales
            extra_files = request.FILES.getlist('extra_imagenes')
            if extra_files:
                import os
                from django.conf import settings
                base_dir = os.path.join(settings.MEDIA_ROOT, 'productos', 'extra', str(producto.id))
                os.makedirs(base_dir, exist_ok=True)
                for f in extra_files:
                    safe_name = f.name
                    dest_path = os.path.join(base_dir, safe_name)
                    if os.path.exists(dest_path):
                        try:
                            os.remove(dest_path)
                        except Exception:
                            pass
                    with open(dest_path, 'wb+') as dest:
                        for chunk in f.chunks():
                            dest.write(chunk)
            # Registrar movimiento de stock inicial si hay cantidad
            if producto.cantidad_disponible > 0:
                Stock.objects.create(
                    producto=producto,
                    tipo='E',
                    cantidad=producto.cantidad_disponible,
                    descripcion='Stock inicial al crear producto'
                )
            messages.success(request, 'Producto creado exitosamente.')
            return redirect('admin_productos')
        else:
            messages.error(request, 'Corrige los errores en el formulario.')
    else:
        form = ProductoForm()
    return render(request, 'vlastef_app/admin_panel/producto_form.html', {'form': form, 'titulo': 'Nuevo Producto'})

# Editar producto
@user_passes_test(staff_required, login_url='index')
def admin_producto_editar_view(request, prod_id):
    producto = get_object_or_404(Producto, id=prod_id)
    cantidad_anterior = producto.cantidad_disponible
    # Marcar borrado de principal dentro del guardado normal
    if request.method == 'POST' and request.POST.get('delete_principal'):
        from django.conf import settings
        import os
        # Eliminar archivo físico si existe
        if producto.imagen and producto.imagen.name:
            abs_path = os.path.join(settings.MEDIA_ROOT, producto.imagen.name)
            if os.path.exists(abs_path):
                try:
                    os.remove(abs_path)
                except Exception:
                    pass
        producto.imagen = None
        producto.save(update_fields=['imagen'])
    # Borrado de imagen extra
    if request.method == 'POST' and request.POST.get('delete_extra'):
        del_name = request.POST.get('file_name')
        if del_name:
            from django.conf import settings
            import os
            rel = del_name.replace('\\','/')
            extra_prefix = f"productos/extra/{producto.id}/"
            # No borrar si es la principal
            if producto.imagen and producto.imagen.name.replace('\\','/') == rel:
                messages.error(request, 'No puedes eliminar la imagen principal desde este botón.')
            else:
                if rel.startswith(extra_prefix):
                    abs_path = os.path.join(settings.MEDIA_ROOT, rel)
                    if os.path.exists(abs_path):
                        try:
                            os.remove(abs_path)
                            messages.success(request, 'Imagen adicional eliminada.')
                        except Exception:
                            messages.error(request, 'No se pudo eliminar la imagen adicional.')
            return redirect('admin_producto_editar', prod_id=producto.id)
    # Procesar cambio de principal mediante POST separado
    if request.method == 'POST' and request.POST.get('set_principal'):
        filename = request.POST.get('file_name')
        if filename:
            from django.conf import settings
            import os, shutil
            rel = filename.replace('\\','/')
            # Esperamos ruta completa relativa dentro de MEDIA_ROOT
            extra_dir = f"productos/extra/{producto.id}/"
            principal_dir = f"productos/{producto.id}/"
            if rel.startswith(extra_dir):
                select_abs = os.path.join(settings.MEDIA_ROOT, rel)
                if os.path.exists(select_abs):
                    # Mover actual principal a extras si existe
                    current = producto.imagen.name.replace('\\','/') if producto.imagen else None
                    if current and current.startswith(principal_dir):
                        curr_abs = os.path.join(settings.MEDIA_ROOT, current)
                        if os.path.exists(curr_abs):
                            os.makedirs(os.path.join(settings.MEDIA_ROOT, extra_dir), exist_ok=True)
                            try:
                                target_extra = os.path.join(settings.MEDIA_ROOT, extra_dir, os.path.basename(curr_abs))
                                if os.path.exists(target_extra):
                                    try:
                                        os.remove(target_extra)
                                    except Exception:
                                        pass
                                shutil.move(curr_abs, target_extra)
                            except Exception:
                                pass
                    # Mover seleccionada a principal
                    os.makedirs(os.path.join(settings.MEDIA_ROOT, principal_dir), exist_ok=True)
                    new_princ_name = os.path.basename(select_abs)
                    try:
                        target_princ = os.path.join(settings.MEDIA_ROOT, principal_dir, new_princ_name)
                        if os.path.exists(target_princ):
                            try:
                                os.remove(target_princ)
                            except Exception:
                                pass
                        shutil.move(select_abs, target_princ)
                        producto.imagen.name = f"{principal_dir}{new_princ_name}"
                        producto.save(update_fields=['imagen'])
                        messages.success(request, 'Imagen principal actualizada.')
                    except Exception:
                        messages.error(request, 'No se pudo cambiar la imagen principal.')
                    return redirect('admin_producto_editar', prod_id=producto.id)
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES, instance=producto)
        if form.is_valid():
            producto = form.save()
            # Reubicar nueva principal si se subió  
            if producto.imagen and producto.imagen.name:
                from django.conf import settings
                import os, shutil
                rel_main = producto.imagen.name.replace('\\','/')
                target_dir = os.path.join(settings.MEDIA_ROOT, 'productos', str(producto.id))
                if rel_main.startswith('productos/') and not rel_main.startswith(f'productos/{producto.id}/'):
                    abs_main = os.path.join(settings.MEDIA_ROOT, rel_main)
                    os.makedirs(target_dir, exist_ok=True)
                    new_name = os.path.basename(abs_main)
                    try:
                        target_princ = os.path.join(target_dir, new_name)
                        if os.path.exists(target_princ):
                            try:
                                os.remove(target_princ)
                            except Exception:
                                pass
                        shutil.move(abs_main, target_princ)
                        producto.imagen.name = f"productos/{producto.id}/{new_name}"
                        producto.save(update_fields=['imagen'])
                    except Exception:
                        pass
            # Guardar nuevas imágenes adicionales
            extra_files = request.FILES.getlist('extra_imagenes')
            if extra_files:
                import os
                from django.conf import settings
                base_dir = os.path.join(settings.MEDIA_ROOT, 'productos', 'extra', str(producto.id))
                os.makedirs(base_dir, exist_ok=True)
                for f in extra_files:
                    safe_name = f.name
                    dest_path = os.path.join(base_dir, safe_name)
                    if os.path.exists(dest_path):
                        try:
                            os.remove(dest_path)
                        except Exception:
                            pass
                    with open(dest_path, 'wb+') as dest:
                        for chunk in f.chunks():
                            dest.write(chunk)
            # Registrar aumento de stock si corresponde
            diferencia = producto.cantidad_disponible - cantidad_anterior
            if diferencia > 0:
                Stock.objects.create(
                    producto=producto,
                    tipo='E',
                    cantidad=diferencia,
                    descripcion='Aumento de stock al editar producto'
                )
            elif diferencia < 0:
                Stock.objects.create(
                    producto=producto,
                    tipo='S',
                    cantidad=abs(diferencia),
                    descripcion='Reducción de stock al editar producto'
                )
            messages.success(request, 'Producto actualizado exitosamente.')
            return redirect('admin_productos')
        else:
            messages.error(request, 'Corrige los errores en el formulario.')
    else:
        form = ProductoForm(instance=producto)
    return render(request, 'vlastef_app/admin_panel/producto_form.html', {'form': form, 'titulo': 'Editar Producto', 'producto': producto})

# Eliminar producto
@user_passes_test(staff_required, login_url='index')
def admin_producto_eliminar_view(request, prod_id):
    producto = get_object_or_404(Producto, id=prod_id)
    # Conservar movimientos de stock de productos eliminados
    Stock.objects.filter(producto=producto).update(producto=None)
    producto.delete()
    messages.success(request, 'Producto eliminado exitosamente.')
    return redirect('admin_productos')


# PROVEEDORES
@user_passes_test(staff_required, login_url='index')
def admin_proveedores_view(request):
    q = request.GET.get('q', '').strip()
    proveedores_qs = Proveedor.objects.all().order_by('-id')
    if q:
        proveedores_qs = proveedores_qs.filter(
            Q(nombre__icontains=q) |
            Q(correo__icontains=q) |
            Q(telefono__icontains=q) |
            Q(direccion__icontains=q)
        )
    total_proveedores = proveedores_qs.count()
    return render(request, 'vlastef_app/admin_panel/proveedores.html', {
        'proveedores': proveedores_qs,
        'query': q,
        'total_proveedores': total_proveedores,
    })

# Detalles del proveedor
@user_passes_test(staff_required, login_url='index')
def admin_proveedor_detalle_view(request, prov_id):
    proveedor = get_object_or_404(Proveedor.objects.prefetch_related('productos'), id=prov_id)
    productos = proveedor.productos.all().select_related('categoria', 'proveedor')
    total_productos = productos.count()
    return render(request, 'vlastef_app/admin_panel/proveedor_detalle.html', {
        'proveedor': proveedor,
        'productos': productos,
        'total_productos': total_productos,
    })

# Crear proveedor
@user_passes_test(staff_required, login_url='index')
def admin_proveedor_crear_view(request):
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Proveedor creado exitosamente.')
            return redirect('admin_proveedores')
        else:
            # Mensaje específico si el nombre ya existe y limpiar el campo
            nombre_errors = form.errors.get('nombre')
            if nombre_errors and any('Ya existe un proveedor con ese nombre.' in str(e) for e in nombre_errors):
                messages.error(request, 'Ya existe un proveedor con ese nombre.')
                mutable = request.POST.copy()
                mutable['nombre'] = ''
                form = ProveedorForm(mutable)
            else:
                correo_errors = form.errors.get('correo')
                if correo_errors and any('Ya existe un proveedor con ese correo.' in str(e) for e in correo_errors):
                    messages.error(request, 'Ya existe un proveedor con ese correo.')
                    mutable = request.POST.copy()
                    mutable['correo'] = ''
                    form = ProveedorForm(mutable)
                else:
                    telefono_errors = form.errors.get('telefono')
                    if telefono_errors and any('Ya existe un proveedor con ese teléfono.' in str(e) for e in telefono_errors):
                        messages.error(request, 'Ya existe un proveedor con ese teléfono.')
                        mutable = request.POST.copy()
                        mutable['telefono'] = ''
                        form = ProveedorForm(mutable)
                    elif form.errors.get('correo'):
                        messages.error(request, 'Ingresa un correo válido.')
                    else:
                        messages.error(request, 'Corrige los errores en el formulario.')
    else:
        form = ProveedorForm()
    return render(request, 'vlastef_app/admin_panel/proveedor_form.html', {'form': form, 'titulo': 'Nuevo Proveedor'})

# Editar proveedor
@user_passes_test(staff_required, login_url='index')
def admin_proveedor_editar_view(request, prov_id):
    proveedor = get_object_or_404(Proveedor, id=prov_id)
    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            form.save()
            messages.success(request, 'Proveedor actualizado exitosamente.')
            return redirect('admin_proveedores')
        else:
            nombre_errors = form.errors.get('nombre')
            if nombre_errors and any('Ya existe un proveedor con ese nombre.' in str(e) for e in nombre_errors):
                messages.error(request, 'Ya existe un proveedor con ese nombre.')
                mutable = request.POST.copy()
                mutable['nombre'] = ''
                form = ProveedorForm(mutable, instance=proveedor)
            else:
                correo_errors = form.errors.get('correo')
                if correo_errors and any('Ya existe un proveedor con ese correo.' in str(e) for e in correo_errors):
                    messages.error(request, 'Ya existe un proveedor con ese correo.')
                    mutable = request.POST.copy()
                    mutable['correo'] = ''
                    form = ProveedorForm(mutable, instance=proveedor)
                else:
                    telefono_errors = form.errors.get('telefono')
                    if telefono_errors and any('Ya existe un proveedor con ese teléfono.' in str(e) for e in telefono_errors):
                        messages.error(request, 'Ya existe un proveedor con ese teléfono.')
                        mutable = request.POST.copy()
                        mutable['telefono'] = ''
                        form = ProveedorForm(mutable, instance=proveedor)
                    elif form.errors.get('correo'):
                        messages.error(request, 'Ingresa un correo válido.')
                    else:
                        messages.error(request, 'Corrige los errores en el formulario.')
    else:
        form = ProveedorForm(instance=proveedor)
    return render(request, 'vlastef_app/admin_panel/proveedor_form.html', {'form': form, 'titulo': 'Editar Proveedor', 'proveedor': proveedor})

# Eliminar proveedor
@user_passes_test(staff_required, login_url='index')
def admin_proveedor_eliminar_view(request, prov_id):
    proveedor = get_object_or_404(Proveedor, id=prov_id)
    proveedor.delete()
    messages.success(request, 'Proveedor eliminado exitosamente.')
    return redirect('admin_proveedores')


# COMENTARIOS
@user_passes_test(staff_required, login_url='index')
def admin_comentarios_view(request):
    q = request.GET.get('q', '').strip()
    comentarios_qs = Comentario.objects.select_related('cliente__usuario', 'producto').order_by('-id')
    if q:
        comentarios_qs = comentarios_qs.filter(
            Q(texto__icontains=q) |
            Q(cliente__usuario__username__icontains=q) |
            Q(producto__nombre__icontains=q)
        )
    total_comentarios = comentarios_qs.count()
    return render(request, 'vlastef_app/admin_panel/comentarios.html', {
        'comentarios': comentarios_qs,
        'query': q,
        'total_comentarios': total_comentarios,
    })

# Eliminar comentario
@user_passes_test(staff_required, login_url='index')
def admin_comentario_eliminar_view(request, com_id):
    comentario = get_object_or_404(Comentario, id=com_id)
    comentario.delete()
    messages.success(request, 'Comentario eliminado exitosamente.')
    return redirect('admin_comentarios')


# VENTAS
@user_passes_test(staff_required, login_url='index')
def admin_ventas_view(request):
    q = request.GET.get('q', '').strip()
    ventas_qs = Venta.objects.select_related('cliente__usuario').order_by('-id')
    
    if q:
        ventas_qs = ventas_qs.filter(
            Q(numero_factura__icontains=q) |
            Q(cliente__usuario__username__icontains=q) |
            Q(cliente__nombres__icontains=q) |
            Q(cliente__apellidos__icontains=q)
        )
    
    total_ventas = ventas_qs.count()
    return render(request, 'vlastef_app/admin_panel/ventas.html', {
        'ventas': ventas_qs,
        'query': q,
        'total_ventas': total_ventas,
    })

# Detalles de la venta
@user_passes_test(staff_required, login_url='index')
def admin_venta_detalle_view(request, venta_id):
    venta = get_object_or_404(Venta.objects.select_related('cliente__usuario'), id=venta_id)
    
    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado')
        if nuevo_estado in dict(Venta.ESTADO_CHOICES):
            # Si se marca como Pagada, validar stock y registrar salida
            if nuevo_estado == 'P':
                detalles = DetalleVenta.objects.filter(venta=venta).select_related('producto')
                # Validar stock suficiente para todos los productos
                insuficientes = []
                for d in detalles:
                    if d.producto.cantidad_disponible < d.cantidad:
                        insuficientes.append((d.producto.nombre, d.producto.cantidad_disponible, d.cantidad))
                if insuficientes:
                    # No cambiar estado si no hay stock
                    pass  # No message
                else:
                    # Descontar stock y registrar movimientos de salida
                    for d in detalles:
                        # Registrar movimiento en historial de Stock
                        Stock.objects.create(
                            producto=d.producto,
                            tipo='S',
                            cantidad=d.cantidad,
                            descripcion=f'Venta #{venta.id}'
                        )
                        # Actualizar cantidad disponible
                        d.producto.cantidad_disponible -= d.cantidad
                        d.producto.save()
                    venta.estado = nuevo_estado
                    venta.save()
                    messages.success(request, 'Estado de la venta actualizado correctamente.')
            else:
                venta.estado = nuevo_estado
                venta.save()
                messages.success(request, 'Estado de la venta actualizado correctamente.')
            return redirect('admin_venta_detalle', venta_id=venta.id)

    detalles = DetalleVenta.objects.filter(venta=venta).select_related('producto')
    total_calculado = detalles.aggregate(Sum('subtotal'))['subtotal__sum'] or 0
    
    return render(request, 'vlastef_app/admin_panel/venta_detalle.html', {
        'venta': venta,
        'detalles': detalles,
        'total_calculado': total_calculado,
    })


# CARRITO
@user_passes_test(staff_required, login_url='index')
def admin_carritos_view(request):
    q = request.GET.get('q', '').strip()
    carritos_qs = Carrito.objects.select_related('cliente__usuario').annotate(items_count=Count('detalles')).order_by('-id')
    
    if q:
        carritos_qs = carritos_qs.filter(
            Q(cliente__usuario__username__icontains=q) |
            Q(cliente__nombres__icontains=q) |
            Q(cliente__apellidos__icontains=q)
        )
        
    total_carritos = carritos_qs.count()
    return render(request, 'vlastef_app/admin_panel/carritos.html', {
        'carritos': carritos_qs,
        'query': q,
        'total_carritos': total_carritos,
    })

# Detalles del carrito
@user_passes_test(staff_required, login_url='index')
def admin_carrito_detalle_view(request, carrito_id):
    carrito = get_object_or_404(Carrito.objects.select_related('cliente__usuario'), id=carrito_id)
    detalles = CarritoDetalle.objects.filter(carrito=carrito).select_related('producto')
    total_carrito = detalles.aggregate(Sum('subtotal'))['subtotal__sum'] or 0
    return render(request, 'vlastef_app/admin_panel/carrito_detalle.html', {
        'carrito': carrito,
        'detalles': detalles,
        'total_carrito': total_carrito,
    })


# STOCK
@user_passes_test(staff_required, login_url='index')
def admin_stock_view(request):
    q = request.GET.get('q', '').strip()
    movimientos = Stock.objects.select_related('producto').order_by('-fecha')
    
    if q:
        movimientos = movimientos.filter(
            Q(producto__nombre__icontains=q) |
            Q(descripcion__icontains=q)
        )
    
    # Obtener resumen de stock actual por producto
    productos_stock = Producto.objects.all().order_by('nombre')
    
    return render(request, 'vlastef_app/admin_panel/stock.html', {
        'movimientos': movimientos,
        'productos_stock': productos_stock,
        'query': q,
    })

# Crear movimiento de stock
@user_passes_test(staff_required, login_url='index')
def admin_stock_crear_view(request):
    if request.method == 'POST':
        form = StockForm(request.POST)
        if form.is_valid():
            producto = form.cleaned_data['producto']
            tipo = form.cleaned_data['tipo']
            cantidad = form.cleaned_data['cantidad']
            descripcion = form.cleaned_data['descripcion']
            
            # Validar stock suficiente para salidas
            if tipo == 'S' and producto.cantidad_disponible < cantidad:
                messages.error(request, f'No hay suficiente stock de {producto.nombre} para realizar esta salida. Disponible: {producto.cantidad_disponible}')
            else:
                # Crear registro de movimiento
                Stock.objects.create(
                    producto=producto,
                    tipo=tipo,
                    cantidad=cantidad,
                    descripcion=descripcion
                )
                
                # Actualizar stock del producto
                if tipo == 'E':
                    producto.cantidad_disponible += cantidad
                else:
                    producto.cantidad_disponible -= cantidad
                producto.save()
                
                messages.success(request, 'Movimiento de stock registrado exitosamente.')
                return redirect('admin_stock')
    else:
        form = StockForm()
    
    return render(request, 'vlastef_app/admin_panel/stock_form.html', {
        'form': form,
        'titulo': 'Registrar Movimiento de Stock'
    })


# CHECKOUT
@login_required
@login_required(login_url='index')
def checkout_view(request):
    try:
        cliente = request.user.cliente
    except Cliente.DoesNotExist:
        return redirect('catalogo_home')

    # Usar 'activo' en lugar de 'procesado'
    carrito = Carrito.objects.filter(cliente=cliente, activo=True).first()
    if not carrito:
        messages.error(request, 'No tienes un carrito activo.')
        return redirect('catalogo_home')
    
    detalles = CarritoDetalle.objects.filter(carrito=carrito).select_related('producto')
    if not detalles:
        messages.error(request, 'Tu carrito está vacío.')
        return redirect('catalogo_home')
    
    # Calcular total
    total = sum(d.subtotal for d in detalles)
    
    return render(request, 'vlastef_app/checkout.html', {
        'detalles': detalles,
        'total': total
    })

@login_required(login_url='index')
def process_payment_view(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    try:
        cliente = request.user.cliente
    except Cliente.DoesNotExist:
        return JsonResponse({'error': 'Usuario no es cliente'}, status=403)

    carrito = Carrito.objects.filter(cliente=cliente, activo=True).first()
    if not carrito:
        return JsonResponse({'error': 'No tienes un carrito activo.'})
    
    detalles = CarritoDetalle.objects.filter(carrito=carrito).select_related('producto')
    if not detalles:
        return JsonResponse({'error': 'Tu carrito está vacío.'})
    
    # Verificar stock suficiente
    for d in detalles:
        if d.cantidad > d.producto.cantidad_disponible:
            return JsonResponse({'error': f'Stock insuficiente para {d.producto.nombre}. Disponible: {d.producto.cantidad_disponible}'})
    
    # Calcular total
    total = sum(d.subtotal for d in detalles)
    
    # Generar número de factura
    # Formato: #01, #02, etc. basado en el conteo total de ventas
    count = Venta.objects.count() + 1
    numero_factura = f"{count:02d}"
    
    # Obtener método de pago del request
    raw_method = request.POST.get('payment_method', 'cash')
    
    # Map frontend method to model choices
    method_map = {
        'cash': 'EF',
        'card': 'TC',
        'paypal': 'OT' # Should not happen due to frontend block, but fallback
    }
    metodo_pago = method_map.get(raw_method, 'EF')
    
    # Determinar estado inicial: Efectivo -> En proceso ('E'), Tarjeta -> Pagada ('P')
    estado_inicial = 'E' if metodo_pago == 'EF' else 'P'

    # Crear Venta
    venta = Venta.objects.create(
        cliente=cliente, 
        total=total,
        numero_factura=numero_factura,
        metodo_pago=metodo_pago,
        estado=estado_inicial
    )
    
    # Crear DetalleVenta, actualizar stock y crear movimientos
    for d in detalles:
        DetalleVenta.objects.create(
            venta=venta,
            producto=d.producto,
            cantidad=d.cantidad,
            precio_unitario=d.producto.precio_venta,
            subtotal=d.subtotal,
            talla=d.talla,
            color=d.color
        )
        
        # Actualizar stock del producto
        d.producto.cantidad_disponible -= d.cantidad
        d.producto.save()
        
        # Crear movimiento de stock
        Stock.objects.create(
            producto=d.producto,
            tipo='S',
            cantidad=d.cantidad,
            descripcion=f'Venta #{venta.id}'
        )
    
    # Vaciar el carrito (mantenerlo activo para reutilizar)
    carrito.detalles.all().delete()
    # carrito.activo = False # NO desactivar, reutilizar el mismo carrito
    # carrito.save()
    
    # Respuesta JSON para manejar el éxito en el frontend
    return JsonResponse({'success': True, 'message': 'Compra realizada con éxito', 'venta_id': venta.id})


# Manejo de errores CSRF
def csrf_failure(request, reason=""):
    try:
        messages.error(request, 'Error de seguridad (CSRF). Por favor recarga la página e inicia sesión de nuevo.')
    except Exception:
        pass
    return redirect('index')


# EXPORT VIEWS

@user_passes_test(staff_required, login_url='index')
def export_ventas_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="ventas.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"

    # Headers
    headers = ['ID', 'Factura', 'Fecha', 'Cliente', 'Método Pago', 'Estado', 'Total']
    ws.append(headers)

    ventas = Venta.objects.select_related('cliente__usuario').order_by('-id')
    
    for venta in ventas:
        ws.append([
            venta.id,
            venta.numero_factura,
            venta.fecha.strftime('%Y-%m-%d %H:%M'),
            str(venta.cliente),
            venta.get_metodo_pago_display(),
            venta.get_estado_display(),
            float(venta.total)
        ])

    wb.save(response)
    return response

@user_passes_test(staff_required, login_url='index')
def export_ventas_pdf(request):
    ventas = Venta.objects.select_related('cliente__usuario').prefetch_related('detalles__producto').order_by('-id')
    template_path = 'vlastef_app/admin_panel/reports/ventas_pdf.html'
    context = {'ventas': ventas}
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="ventas.pdf"'
    template = get_template(template_path)
    html = template.render(context)
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response

@user_passes_test(staff_required, login_url='index')
def export_stock_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="stock.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos de Stock"

    headers = ['ID', 'Fecha', 'Producto', 'Tipo', 'Cantidad', 'Descripción']
    ws.append(headers)

    movimientos = Stock.objects.select_related('producto').order_by('-fecha')
    
    for mov in movimientos:
        ws.append([
            mov.id,
            mov.fecha.strftime('%Y-%m-%d %H:%M'),
            mov.producto.nombre if mov.producto else 'Producto Eliminado',
            mov.get_tipo_display(),
            mov.cantidad,
            mov.descripcion
        ])

    wb.save(response)
    return response

@user_passes_test(staff_required, login_url='index')
def export_stock_pdf(request):
    movimientos = Stock.objects.select_related('producto').order_by('-fecha')
    template_path = 'vlastef_app/admin_panel/reports/stock_pdf.html'
    context = {'movimientos': movimientos}
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="stock.pdf"'
    template = get_template(template_path)
    html = template.render(context)
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response

@user_passes_test(staff_required, login_url='index')
def export_productos_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="productos.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    headers = ['ID', 'Nombre', 'Categoría', 'Proveedor', 'Precio Compra', 'Precio Venta', 'Stock', 'Descripción']
    ws.append(headers)

    productos = Producto.objects.select_related('categoria', 'proveedor').order_by('-id')
    
    for prod in productos:
        ws.append([
            prod.id,
            prod.nombre,
            prod.categoria.nombre if prod.categoria else '-',
            prod.proveedor.nombre if prod.proveedor else '-',
            float(prod.precio_real),
            float(prod.precio_venta),
            prod.cantidad_disponible,
            prod.descripcion
        ])

    wb.save(response)
    return response

@user_passes_test(staff_required, login_url='index')
def export_productos_pdf(request):
    productos = Producto.objects.select_related('categoria', 'proveedor').order_by('-id')
    template_path = 'vlastef_app/admin_panel/reports/productos_pdf.html'
    context = {'productos': productos}
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="productos.pdf"'
    template = get_template(template_path)
    html = template.render(context)
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response

def link_callback(uri, rel):
    """
    Convert HTML URIs to absolute system paths so xhtml2pdf can access those
    resources
    """
    sUrl = settings.STATIC_URL        # Typically /static/
    mUrl = settings.MEDIA_URL         # Typically /media/
    mRoot = settings.MEDIA_ROOT       # Typically /home/userX/project_static/media/

    if uri.startswith(mUrl):
        path = os.path.join(mRoot, uri.replace(mUrl, ""))
    elif uri.startswith(sUrl):
        sUri = uri.replace(sUrl, "")
        # First try to find it using finders (works in dev)
        path = finders.find(sUri)
        if not path:
            # If not found by finders, try STATIC_ROOT (works in prod if collectstatic run)
            sRoot = getattr(settings, 'STATIC_ROOT', None)
            if sRoot:
                path = os.path.join(sRoot, sUri)
    else:
        return uri

    # make sure that file exists
    if not path or not os.path.isfile(path):
        return None
    return path

@login_required
def download_invoice_view(request, venta_id):
    venta = get_object_or_404(Venta, id=venta_id)
    # Check permission: user must be the client or admin
    if not request.user.is_staff and venta.cliente.usuario != request.user:
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied

    template_path = 'vlastef_app/factura_pdf.html'
    context = {
        'venta': venta,
    }
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="factura_{venta.numero_factura}.pdf"'
    template = get_template(template_path)
    html = template.render(context)
    pisa_status = pisa.CreatePDF(html, dest=response, link_callback=link_callback)
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response
